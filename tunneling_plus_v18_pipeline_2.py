#!/usr/bin/env python3
"""
Tunneling+ v18 Pipeline — Statcast Edition
===========================================
IDENTICAL model to v17. Replaces TJ Stats CSV input with live Statcast pull.

Calibration fixes applied to match TJ Stats aggregation:
  1. hb sign:        hand-dependent. RHP: pfx_x*12*-1. LHP: pfx_x*12.
                     pfx_x*12*-1 gives correct arm-side hb for RHP.
                     LHP requires no flip (pfx_x*12) to match TJ Stats convention.
  2. All pitches:    no runner/windup filter — TJ Stats uses all pitches; windup-only
                     shifts hb averages enough to push SL/CU td below the 0.5" floor
                     in score_pair_metrics(), silently dropping the pair.
  3. release_side:   corrected with extension-based formula derived from diagnostic:
                     rs_corrected = release_pos_x + (0.3655 * extension - 2.4608)
  4. Min pitches:    10 per pitch type (same as V17), valid pitch types only

USAGE:
    python3 tunneling_plus_v18_pipeline.py \\
        --start 2026-03-27 \\
        --end   2026-05-03 \\
        --output tunneling_plus_2026_v18.xlsx \\
        --json   results_v18.json

REQUIREMENTS:
    pip install pybaseball pandas numpy openpyxl
"""

import argparse, sys
import pandas as pd
import numpy as np
import json
from itertools import combinations
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Model constants (identical to V17) ───────────────────────────────────────
TUNNEL_FT  = 23.8
LG_AVG_PD  = 20.07
TD_TUNNEL  = 6.6
PD_SPEED   = 14.9
VD_SPEED   = 6.8
MIN_RATIO  = 1.5

# Valid pitch types (excludes PO, IN, FA, automatic ball etc.)
VALID_TYPES = {
    'FF','SI','FC','FS','FT',
    'CH','SC',
    'SL','ST',
    'CU','KC','KN',
    'EP','FO','SV','CS'
}

# release_side correction coefficients (fitted from diagnostic comparison
# of Statcast raw vs TJ Stats values across FF/SL/CU for Jacob Misiorowski)
# correction = RS_A * extension + RS_B
RS_A =  0.3655
RS_B = -2.4608


# ── Model functions (identical to V17) ───────────────────────────────────────
def axis_tunnel_score(dx, dz):
    ew = 1.0 / (1.0 + abs(dx))
    ns = 1.0 / (1.0 + abs(dz))
    base = max(ew, ns)
    return base + max(np.sqrt(ew * ns) - base * 0.7, 0)


def score_pair_metrics(r1, r2):
    dx_tun = r1['tunnel_x'] - r2['tunnel_x']
    dz_tun = r1['tunnel_z'] - r2['tunnel_z']
    dx_plt = r1['plate_x']  - r2['plate_x']
    dz_plt = r1['plate_z']  - r2['plate_z']
    rd = np.sqrt((r1['rel_x'] - r2['rel_x'])**2 + (r1['rel_z'] - r2['rel_z'])**2)
    td  = np.sqrt(dx_tun**2 + dz_tun**2)
    pd_ = np.sqrt(dx_plt**2 + dz_plt**2)
    if td < 0.5:
        return None
    vd    = abs(r1['velo'] - r2['velo'])
    ratio = pd_ / td
    uw    = np.sqrt(r1['pitch_frac'] * r2['pitch_frac'])
    btr   = max(pd_ - td, 0) / td
    rtr   = rd / td if td > 0 else 0
    ax    = axis_tunnel_score(dx_tun, dz_tun)
    ix    = ax * (pd_ / LG_AVG_PD)
    er    = td / pd_ if pd_ > 0.5 else 1.0
    lbm   = 1.0 - 0.5 * er
    qw    = uw * ratio
    return dict(tr=ratio, btr=btr, rtr=rtr, ix=ix, lbm=lbm, vd=vd,
                uw=uw, qw=qw, td=td, pd=pd_, tm=vd * (1.0 / (1.0 + pd_)))


def compute_tunnel_composite(tunnel_pairs):
    tw = sum(p['qw'] for p in tunnel_pairs)
    if tw == 0:
        return 0, 0
    def w(k): return sum(p[k] * p['qw'] for p in tunnel_pairs) / tw
    tc = (0.35 * w('tr') + 0.20 * w('btr') + 0.15 * w('ix') +
          0.15 * (-w('rtr')) + 0.10 * w('lbm') + 0.05 * w('vd'))
    avg_tr = w('tr')
    return tc, avg_tr


# ── Statcast data loading ─────────────────────────────────────────────────────
def load_statcast(start_date, end_date, verbose=True):
    """
    Pull raw Statcast, aggregate per pitcher/pitch-type,
    apply all calibration fixes, compute geometry — returns
    the same df format as V17's load_data().
    """
    from pybaseball import statcast
    from pybaseball import cache as pb_cache
    pb_cache.enable()

    if verbose:
        print(f'  Fetching Statcast {start_date} → {end_date}...')
    raw = statcast(start_date, end_date)

    if verbose:
        print(f'  Raw rows: {len(raw):,}')

    # ── Filter 1: valid pitch types ───────────────────────────────────────────
    raw = raw[raw['pitch_type'].isin(VALID_TYPES)].copy()

    # ── Filter 2: non-null movement and release data ──────────────────────────
    required = ['pfx_x','pfx_z','release_pos_x','release_pos_z',
                'release_extension','release_speed']
    raw = raw[raw[required].notna().all(axis=1)]

    # ── Filter 3: sane velocity ───────────────────────────────────────────────
    raw = raw[raw['release_speed'].between(50, 110)]

    # NOTE: No windup filter applied. TJ Stats uses all pitches regardless of
    # runners on base. Windup-only shifts hb/ivb averages enough to push
    # SL/CU tunnel distance below the 0.5" floor, silently dropping the pair.

    # ── Build hand map first (needed for hb sign) ───────────────────────────
    hand_map = {}
    if 'p_throws' in raw.columns:
        for pid, grp in raw.groupby('pitcher'):
            h = grp['p_throws'].mode()
            if len(h):
                hand_map[int(pid)] = h.iloc[0]

    # ── Compute hb with hand-dependent sign fix ───────────────────────────────
    # Statcast pfx_x * 12 * -1 gives correct arm-side hb for RHP.
    # For LHP the same formula inverts the sign — so LHP needs pfx_x * 12 (no flip).
    # Result: positive hb = arm side for that pitcher, matching TJ Stats convention.
    if 'p_throws' in raw.columns:
        raw['hb'] = np.where(
            raw['p_throws'] == 'L',
            raw['pfx_x'] * 12,          # LHP: no flip
            raw['pfx_x'] * 12 * -1      # RHP: flip
        )
    else:
        raw['hb'] = raw['pfx_x'] * 12 * -1  # default RHP convention
    raw['ivb'] = raw['pfx_z'] * 12

    # ── Aggregate per pitcher/pitch-type ─────────────────────────────────────

    # Team: most recent game's fielding team
    team_map = {}
    if 'home_team' in raw.columns and 'inning_topbot' in raw.columns:
        raw['_pitcher_team'] = np.where(
            raw['inning_topbot'] == 'Top',
            raw['home_team'],
            raw['away_team'] if 'away_team' in raw.columns else raw['home_team']
        )
        for pid, grp in raw.groupby('pitcher'):
            team_map[int(pid)] = grp['_pitcher_team'].iloc[-1]

    # Name map
    name_map = {}
    if 'player_name' in raw.columns:
        for pid, grp in raw.groupby('pitcher'):
            nm = grp['player_name'].iloc[0]
            # Statcast gives "Last, First" — flip to "First Last"
            if isinstance(nm, str) and ',' in nm:
                parts = nm.split(',', 1)
                nm = parts[1].strip() + ' ' + parts[0].strip()
            name_map[int(pid)] = nm

    agg = raw.groupby(['pitcher','pitch_type']).agg(
        n               = ('release_speed','count'),
        velo            = ('release_speed','mean'),
        ivb             = ('ivb','mean'),
        hb              = ('hb','mean'),
        extension       = ('release_extension','mean'),
        release_height  = ('release_pos_z','mean'),
        release_side_raw= ('release_pos_x','mean'),
    ).reset_index()

    # ── Apply release_side correction ────────────────────────────────────────
    # Derived from diagnostic: correction = RS_A*ext + RS_B
    # shifts Statcast release_pos_x to match TJ Stats release_side
    agg['release_side'] = agg['release_side_raw'] + (RS_A * agg['extension'] + RS_B)

    # ── Filter: min 10 pitches per pitch type ─────────────────────────────────
    agg = agg[agg['n'] >= 10].copy()
    agg.rename(columns={'pitcher':'pitcher_id', 'n':'pitches'}, inplace=True)

    # ── Attach metadata ───────────────────────────────────────────────────────
    agg['pitcher_name'] = agg['pitcher_id'].map(name_map).fillna('Unknown')
    agg['hand']         = agg['pitcher_id'].map(hand_map)
    agg['pitcher_team'] = agg['pitcher_id'].map(team_map).fillna('???')
    agg['tj_stuff_plus'] = 100.0   # not available from Statcast; placeholder

    # ── Geometry (identical to V17 load_data) ────────────────────────────────
    agg['release_dist'] = 60.5 - agg['extension']
    agg['mv_scale']     = ((agg['release_dist'] - TUNNEL_FT) / agg['release_dist']) ** 2
    agg['tunnel_x']     = agg['release_side'] * 12 + agg['hb'] * agg['mv_scale']
    agg['tunnel_z']     = agg['release_height'] * 12 + agg['ivb'] * agg['mv_scale']
    agg['plate_x']      = agg['release_side'] * 12 + agg['hb']
    agg['plate_z']      = agg['release_height'] * 12 + agg['ivb']
    agg['rel_x']        = agg['release_side'] * 12
    agg['rel_z']        = agg['release_height'] * 12

    # Fill missing hand from release_side heuristic
    for pid, grp in agg[agg['hand'].isna()].groupby('pitcher_id'):
        ff = grp[grp['pitch_type'].isin(['FF','SI'])]
        ref = ff.iloc[0] if len(ff) else grp.iloc[0]
        agg.loc[agg['pitcher_id'] == pid, 'hand'] = (
            'R' if ref['release_side'] < 0 else 'L')

    agg['pitch_frac'] = agg.groupby('pitcher_id')['pitches'].transform(
        lambda x: x / x.sum())

    if verbose:
        print(f'  Aggregated: {len(agg)} pitch-type rows, '
              f'{agg["pitcher_id"].nunique()} pitchers')

    return agg


# ── Model runner (identical to V17) ──────────────────────────────────────────
def run_model(df):
    results_classified = []
    results_fallback   = []

    for pitcher_id, group in df.groupby('pitcher_id'):
        if len(group) < 2:
            continue
        pname  = group['pitcher_name'].iloc[0]
        team   = group['pitcher_team'].iloc[0]
        hand   = group['hand'].iloc[0]
        total  = int(group['pitches'].sum())
        rows   = group.to_dict('records')

        tunnel_pairs = []
        speed_pairs  = []
        all_pairs    = []
        n_irrel      = 0

        for r1, r2 in combinations(rows, 2):
            m = score_pair_metrics(r1, r2)
            if m is None:
                continue
            all_pairs.append(m)
            is_t = m['td'] < TD_TUNNEL and m['tr'] > MIN_RATIO
            is_s = m['pd'] < PD_SPEED  and m['vd'] > VD_SPEED
            if is_t:
                tunnel_pairs.append(m)
            elif is_s:
                speed_pairs.append(m)
            else:
                n_irrel += 1

        pitch_details = [
            {'type': r['pitch_type'], 'pitches': int(r['pitches']),
             'velo': round(r['velo'], 1), 'ivb': round(r['ivb'], 1),
             'hb': round(r['hb'], 1), 'stuff': 100.0,
             'frac': round(r['pitch_frac'] * 100, 1)}
            for _, r in group.iterrows()
        ]

        base = dict(pitcher_id=int(pitcher_id), name=pname, team=team, hand=hand,
                    pitches=total, n_types=len(group), pitch_details=pitch_details)

        if tunnel_pairs or speed_pairs:
            tc, avg_tr = compute_tunnel_composite(tunnel_pairs) if tunnel_pairs else (0, 0)
            temporal = 0
            if speed_pairs:
                sw = sum(p['uw'] for p in speed_pairs)
                temporal = sum(p['tm'] * p['uw'] for p in speed_pairs) / sw
            results_classified.append({**base,
                'n_tunnel_pairs': len(tunnel_pairs),
                'n_speed_pairs':  len(speed_pairs),
                'n_irrelevant':   n_irrel,
                'fallback': False,
                'tunnel_composite': round(tc, 4),
                'avg_tunnel_ratio': round(avg_tr, 3),
                'temporal': round(temporal, 4)})

        elif all_pairs:
            tc, avg_tr = compute_tunnel_composite(all_pairs)
            results_fallback.append({**base,
                'n_tunnel_pairs': 0, 'n_speed_pairs': 0, 'n_irrelevant': 0,
                'fallback': True,
                'tunnel_composite': round(tc, 4),
                'avg_tunnel_ratio': round(avg_tr, 3),
                'temporal': 0.0})

    return results_classified, results_fallback


# ── Normalization (identical to V17) ─────────────────────────────────────────
def normalize(results_classified, results_fallback):
    rdf = pd.DataFrame(results_classified)
    norm_params = {}
    comp_params = {}

    for hand in ['R', 'L']:
        mask = rdf['hand'] == hand
        sub_tc = rdf.loc[mask, 'tunnel_composite']
        sub_tm = rdf.loc[mask, 'temporal']
        mu_tc, sd_tc = sub_tc.mean(), sub_tc.std()
        mu_tm, sd_tm = sub_tm.mean(), sub_tm.std()
        rdf.loc[mask, 'z_tunnel']   = (sub_tc - mu_tc) / sd_tc
        rdf.loc[mask, 'z_temporal'] = (sub_tm - mu_tm) / sd_tm
        norm_params[hand] = dict(mu_tc=mu_tc, sd_tc=sd_tc, mu_tm=mu_tm, sd_tm=sd_tm)

    rdf['composite'] = 0.80 * rdf['z_tunnel'] + 0.20 * rdf['z_temporal']

    for hand in ['R', 'L']:
        mask = rdf['hand'] == hand
        sub  = rdf.loc[mask, 'composite']
        mu_c, sd_c = sub.mean(), sub.std()
        rdf.loc[mask, 'tunneling_plus'] = (
            (100 + ((sub - mu_c) / sd_c) * 15).clip(60, 160).round(1))
        comp_params[hand] = dict(mu_c=mu_c, sd_c=sd_c)

    rdf_f = pd.DataFrame(results_fallback) if results_fallback else pd.DataFrame()
    if not rdf_f.empty:
        for hand in ['R', 'L']:
            mask = rdf_f['hand'] == hand
            if not mask.any():
                continue
            np_ = norm_params[hand]
            cp  = comp_params[hand]
            z_tc = (rdf_f.loc[mask, 'tunnel_composite'] - np_['mu_tc']) / np_['sd_tc']
            z_tm = (rdf_f.loc[mask, 'temporal']          - np_['mu_tm']) / np_['sd_tm']
            comp = 0.80 * z_tc + 0.20 * z_tm
            rdf_f.loc[mask, 'tunneling_plus'] = (
                (100 + ((comp - cp['mu_c']) / cp['sd_c']) * 15).clip(60, 160).round(1))

    rdf_all = pd.concat([rdf, rdf_f], ignore_index=True) if not rdf_f.empty else rdf
    q = rdf_all[rdf_all['pitches'] >= 50].copy()
    q['rank'] = q['tunneling_plus'].rank(ascending=False, method='min').astype(int)

    def last_name(n): return n.split()[-1].lower()
    q = q.sort_values(['rank', 'name'],
                      key=lambda c: c.map(last_name) if c.name == 'name' else c
                      ).reset_index(drop=True)
    return q


# ── Excel export (identical to V17, version label updated) ───────────────────
def build_excel(q, output_path, start_date, end_date):
    wb = Workbook()

    HDR_FONT  = Font(name='Arial', bold=True, color='FFFFFF', size=10)
    HDR_FILL  = PatternFill('solid', start_color='1F3864')
    HDR_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
    CTR       = Alignment(horizontal='center', vertical='center')
    thin      = Side(style='thin', color='CCCCCC')
    BDR       = Border(left=thin, right=thin, top=thin, bottom=thin)

    def tier_fill(tp):
        if tp >= 130: return PatternFill('solid', start_color='DDEEFF')
        if tp >= 115: return PatternFill('solid', start_color='E8F5E9')
        if tp >= 90:  return PatternFill('solid', start_color='FFFFFF')
        if tp >= 80:  return PatternFill('solid', start_color='FFF8DD')
        return              PatternFill('solid', start_color='FFE8E8')

    def tier_color(tp):
        if tp >= 130: return '185FA5'
        if tp >= 115: return '3B6D11'
        if tp >= 90:  return '000000'
        if tp >= 80:  return '854F0B'
        return              'A32D2D'

    def hdr_row(ws, headers, widths):
        ws.row_dimensions[1].height = 36
        for c, (h, w) in enumerate(zip(headers, widths), 1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.font = HDR_FONT; cell.fill = HDR_FILL
            cell.alignment = HDR_ALIGN; cell.border = BDR
            ws.column_dimensions[get_column_letter(c)].width = w

    def dcell(ws, row, col, val, fmt=None, bold=False, color='000000', fill=None):
        cell = ws.cell(row=row, column=col, value=val)
        cell.font = Font(name='Arial', size=10, bold=bold, color=color)
        cell.alignment = CTR; cell.border = BDR
        if fmt:  cell.number_format = fmt
        if fill: cell.fill = fill

    records = q.drop(columns=['pitcher_id', 'z_tunnel', 'z_temporal',
                               'composite', 'pitch_details'],
                      errors='ignore').to_dict('records')

    # Sheet 1: Leaderboard
    ws1 = wb.active; ws1.title = 'Tunneling+ Leaderboard'; ws1.freeze_panes = 'A2'
    h1 = ['Rank','Pitcher','Team','Hand','Pitches','Pitch Types',
          'Tunneling+','Tunnel Pairs','Avg Tunnel Ratio',
          'Speed-Change Pairs','Irrelevant Pairs','Temporal Score','Note']
    w1 = [6,22,6,6,8,11,12,13,17,19,16,14,14]
    hdr_row(ws1, h1, w1)
    for r, p in enumerate(records, 2):
        tp = p['tunneling_plus']; tf = tier_fill(tp)
        note = 'Unclassified pairs' if p.get('fallback') else ''
        vals = [p['rank'], p['name'], p['team'], p['hand'], p['pitches'],
                p['n_types'], tp, p['n_tunnel_pairs'], p['avg_tunnel_ratio'],
                p['n_speed_pairs'], p['n_irrelevant'], p['temporal'], note]
        fmts = [None,None,None,None,None,None,'0.0',None,'0.000',
                None,None,'0.0000',None]
        for c, (v, f) in enumerate(zip(vals, fmts), 1):
            color = (tier_color(tp) if c == 7 else
                     '888888' if c == 13 else '000000')
            dcell(ws1, r, c, v, fmt=f, bold=(c==7), color=color, fill=tf)
    ws1.auto_filter.ref = f'A1:{get_column_letter(len(h1))}1'

    # Sheet 2: Pitch Arsenal
    ws2 = wb.create_sheet('Pitch Arsenal'); ws2.freeze_panes = 'A2'
    h2 = ['Rank','Pitcher','Team','Hand','Pitch Type',
          'Usage %','Velocity','IVB (in)','HB (in)']
    w2 = [6,22,6,6,11,9,10,10,10]
    hdr_row(ws2, h2, w2)
    r2 = 2
    for p_rec in q.to_dict('records'):
        details = p_rec.get('pitch_details', [])
        for pd_ in sorted(details, key=lambda x: -x['frac']):
            vals = [p_rec['rank'], p_rec['name'], p_rec['team'], p_rec['hand'],
                    pd_['type'], pd_['frac'], pd_['velo'],
                    pd_['ivb'], pd_['hb']]
            fmts = [None,None,None,None,None,'0.0','0.0','0.0','0.0']
            for c, (v, f) in enumerate(zip(vals, fmts), 1):
                dcell(ws2, r2, c, v, fmt=f)
            r2 += 1
    ws2.auto_filter.ref = f'A1:{get_column_letter(len(h2))}1'

    # Sheet 3: Methodology
    ws3 = wb.create_sheet('Methodology')
    ws3.column_dimensions['A'].width = 28
    ws3.column_dimensions['B'].width = 86
    for col, val in [(1,'Concept'), (2,'Description')]:
        cell = ws3.cell(row=1, column=col, value=val)
        cell.font = HDR_FONT; cell.fill = HDR_FILL
        cell.alignment = HDR_ALIGN; cell.border = BDR
    ws3.row_dimensions[1].height = 24

    methodology = [
        ('Model version',
         f'Tunneling+ v18 — Statcast Edition. '
         f'Identical model to v17. Data: {start_date} to {end_date}. All pitches (no windup filter).'),
        ('Tunnel pair',
         'td < 6.6" (league p50) AND ratio > 1.5x. '
         'Scored on spatial metrics weighted by usage × tunnel_ratio.'),
        ('Speed-change pair',
         'pd < 14.9" (league p33) AND vd > 6.8 mph (league p50). '
         'Scored on temporal: velo_diff × 1/(1+plate_diff).'),
        ('Irrelevant pair',
         'Neither condition met. Excluded for pitchers with other qualifying pairs.'),
        ('Fallback scoring',
         'Pitchers whose pairs all fall outside both classifications are scored on '
         'raw pair metrics. Flagged "Unclassified pairs" in Note column.'),
        ('Quality weighting',
         'Tunnel pairs weighted by usage × tunnel_ratio. '
         'Elite pairs dominate; weak pairs contribute proportionally less.'),
        ('Composite',
         '80% tunnel composite + 20% temporal. '
         'Normalized within handedness to 100=avg, std~15. '
         'Winsorized [60,160]. Min 10 pitches/type, 50 total.'),
        ('Tunnel composite weights',
         '35% Tunnel Ratio + 20% Break:Tunnel + 15% Interaction + '
         '15% Release:Tunnel (inv) + 10% Late Break Mult + 5% Velo Diff'),
        ('Data source',     'Baseball Savant (Statcast). All pitches included — no runner/windup filter.'),
        ('Calibration',
         'hb sign is hand-dependent: RHP = pfx_x*12*-1, LHP = pfx_x*12 (no flip). '
         f'release_side = release_pos_x + ({RS_A}*ext + ({RS_B:.4f})) '
         '(extension-based correction to match TJ Stats reference frame).'),
        ('Sort order',    'Ties broken alphabetically by last name.'),
        ('Constants',
         f'LG_AVG_PD={LG_AVG_PD}", TD_TUNNEL={TD_TUNNEL}", '
         f'PD_SPEED={PD_SPEED}", VD_SPEED={VD_SPEED}mph, MIN_RATIO={MIN_RATIO}x'),
    ]
    for i, (label, desc) in enumerate(methodology, 2):
        for col, val in [(1, label), (2, desc)]:
            cell = ws3.cell(row=i, column=col, value=val)
            cell.font = Font(name='Arial', size=10, bold=(col==1))
            cell.alignment = Alignment(vertical='top', wrap_text=True)
            cell.border = BDR
        ws3.row_dimensions[i].height = 44

    wb.save(output_path)
    print(f'Saved {output_path}')


# ── Entry point ───────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(
        description='Tunneling+ v18 Pipeline (Statcast Edition)',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__)
    parser.add_argument('--start', default='2026-03-27',
                        help='Season start date (YYYY-MM-DD)')
    parser.add_argument('--end',   default=None,
                        help='End date (YYYY-MM-DD), defaults to today')
    parser.add_argument('--output', required=True,
                        help='Output Excel path (.xlsx)')
    parser.add_argument('--json',   default=None,
                        help='Also save results as JSON (optional)')
    args = parser.parse_args()

    from datetime import date
    end_date = args.end or date.today().strftime('%Y-%m-%d')

    print('Loading Statcast data...')
    df = load_statcast(args.start, end_date)

    print('Running model...')
    classified, fallback = run_model(df)
    print(f'  {len(classified)} classified, {len(fallback)} fallback')

    print('Normalizing...')
    q = normalize(classified, fallback)
    print(f'  {len(q)} qualified (50+ pitches)')
    print(f'  Mean T+: {q["tunneling_plus"].mean():.1f}  '
          f'Std: {q["tunneling_plus"].std():.1f}')

    print('Building Excel...')
    build_excel(q, args.output, args.start, end_date)

    if args.json:
        out = q.drop(columns=['pitcher_id','z_tunnel','z_temporal',
                               'composite'], errors='ignore').to_dict('records')
        json.dump(out, open(args.json, 'w'), separators=(',',':'))
        print(f'Saved {args.json}')

    print('Done.')
    print(f'\nTop 10:')
    for _, r in q.head(10).iterrows():
        print(f'  #{int(r["rank"]):<4} {r["name"]:<24} {r["tunneling_plus"]:.1f}')

    # Spot-check Misiorowski
    mis = q[q['name'].str.contains('Misiorowski', na=False)]
    if not mis.empty:
        m = mis.iloc[0]
        print(f'\nMisiorowski: T+={m["tunneling_plus"]}  '
              f'rank=#{int(m["rank"])}  '
              f'tp={m["n_tunnel_pairs"]}  '
              f'ratio={m["avg_tunnel_ratio"]:.3f}')


if __name__ == '__main__':
    main()
